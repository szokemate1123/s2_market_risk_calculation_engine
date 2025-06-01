import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import xlwings as xw
import logging
import time
import re

# Logging beállítása
logging.basicConfig(level=logging.ERROR, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)
file_handler = logging.FileHandler('base_functions.log')
file_handler.setLevel(logging.ERROR)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(file_handler)

def file_exists(file_path):
    """Megnézi, hogy a megadott helyen van-e file és az alapján add visszajelzést."""
    if not os.path.exists(file_path):
        logger.error(f"Ez a file nem elérhető: '{file_path}'")
        raise FileNotFoundError(f"Ez a file nem elérhető: '{file_path}'")
    return True

def get_range_string(wb, range_name):
    """
    Meghatározza a bemenetnek az excel tartományát.
    Azért kell, mert ha üres cella van a range_name-ben, akkor hibát dob az openpyxl mindenképp.
    
    Args:
        wb: OpenPyXL Workbook objektum.
        range_name (str): Named range vagy explicit range (pl. "Sheet1!A1:B3").
    
    Returns:
        str: A stringet adja vissza, amit utána már be tudunk olvasni.
    """
    # Ha explicit tartományt adunk meg, akkor nincs dolgunk
    if '!' in range_name:
        return range_name

    # Named range esetén
    named_range = wb.defined_names.get(range_name)
    if not named_range:
        logger.error(f"Named range '{range_name}' nem található a file-ban.")
        raise ValueError(f"Named range '{range_name}' nem található a file-ban.")
    return named_range.attr_text

def extract_cells(wb, range_string):
    """
    A már explicit tartományban lévő adatokat olvassa be az excel file-ból.
    
    Args:
        wb: OpenPyXL Workbook objektum.
        range_string (str): A tartomány string (pl. "Sheet1!A1:B3,Sheet2!C1:D2").
    
    Returns:
        list: Visszaadja egy listában az összes adatot, amit utána lehet alakítani.
    """
    combined_data = []

    # Ha több tartomány is meg van adva vesszővel elválasztva, akkor azt is kezelje
    for ref in range_string.split(','):
        ws_name, cell_range = ref.split('!')
        ws_name = ws_name.strip("'")

        # Megnézzük, hogy létezik-e a fül a munkafüzetben
        if ws_name not in wb.sheetnames:
            logger.error(f"A '{ws_name}' nevű fül nem létezik.")
            raise ValueError(f"A '{ws_name}' nevű fül nem létezik.")

        ws = wb[ws_name]
        cells = ws[cell_range]
        combined_data.extend([[cell.value for cell in row] for row in cells])

    return combined_data

def dataframe_from_xlsx(file_path: str, range_name: str, header=True) -> pd.DataFrame:
    """
    Ez a beolvasás vége, használja az eddigi függvényeket és a bemeneti tartományt beolvassa dataframe-ként.
    
    Args:
        file_path (str): Az Excel file elérési útja.
        range_name (str): Named range vagy explicit range (pl. "Sheet1!A1:B3").
        header (Bool): Fejléc használata vagy sem.
    
    Returns:
        pd.DataFrame: Egy dataframe, amiben benne van az összes adat amit kértünk.
    """
    # Megnézzük a file működik-e
    file_exists(file_path)

    # A munkafüzet betöltése openpyxl-el (xlsb nem támogatott)
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    # Megtalálja a range_stringet
    range_string = get_range_string(wb, range_name)

    # Adat kinyerése
    data = extract_cells(wb, range_string)

    # Dataframe visszaadása, vagy fejléccel, vagy anélkül
    if header and data:
        return pd.DataFrame(data[1:], columns=data[0])
    return pd.DataFrame(data)

def transpose_sheets(input_file, sheets, output_file):
    """
    A megadott Excel file füljeinek transzponálása és mentése egy új Excel file-ba.
    
    Args:
        input_file (str): Az input Excel file elérési útja.
        sheets (list): A transzponálandó fülek listája.
        output_file (str): Az új Excel file elérési útja.
    """
    try:
        # Pandas Excel writer létrehozása openpyxl motorral
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Minden megadott fül feldolgozása
            for sheet_name in sheets:
                # A fül beolvasása DataFrame-be
                df = pd.read_excel(input_file, sheet_name=sheet_name)
                
                # A DataFrame transzponálása
                transposed_df = df.transpose()
                
                # Az első sor eltávolítása, ha az fejléccé válik a transzponálás után
                transposed_df = transposed_df.iloc[1:, :]
                
                # A transzponált DataFrame írása az új file-ba
                transposed_df.to_excel(writer, sheet_name=sheet_name, header=False, index=False)
    except Exception as e:
        logger.error(f"Hiba a fülek transzponálása közben: {e}")
        raise

def recalculate_file_with_xlwings(file_path):
    """
    Meg kell adni egy file elérési útvonalát
    és xlwings-el teljesen újrakalkulálja a file-t.
    
    Args:
        file_path (str): Az Excel file elérési útja.
    """
    app = None
    try:
        # Megnyitás
        app = xw.App(add_book=False, visible=False)
        wb = app.books.open(file_path)
        
        # Újra kalkulálás
        wb.app.calculate()
        time.sleep(0.5)
        
        # Mentés
        wb.save()
    except Exception as e:
        logger.error(f"Hiba a file újrakalkulálása közben: {e}")
        raise
    finally:
        if app:
            wb.close()
            app.quit()

def adjust_column_width(sheet):
    """
    Az oszlopok szélességének beállítása a maximális cellahossz alapján.
    """
    for column in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                # Ellenőrizzük a cella értékének hosszát
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        # Beállítjuk az oszlop szélességét a maximális hossz alapján
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column_letter].width = adjusted_width

def add_filter_and_paint(sheet):
    """
    Szűrő hozzáadása az első sorhoz és az oszlopok sötétkékre festése.
    """
    # Szűrő hozzáadása az első sorhoz
    sheet.auto_filter.ref = sheet.dimensions

    # Az első sor sötétkékre festése
    dark_blue_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    for cell in sheet[1]:
        cell.fill = dark_blue_fill
        cell.font = cell.font.copy(color="FFFFFF")  # Betűszín beállítása fehérre

def name_ranges(sheet, range_name):
    """
    Tartományok elnevezése az adott lapon.
    
    Args:
    - sheet (Worksheet): Az Excel munkalap.
    - range_name (str): A tartomány neve.
    """
    # Meghatározzuk a tartományt (az összes adat a munkalapon)
    data_range = f"{sheet.dimensions}"
    sheet.parent.create_named_range(range_name, sheet, data_range)

def process_workbook(file_path, range_names):
    """
    Excel munkafüzet feldolgozása: oszlopszélesség beállítása, szűrő hozzáadása, oszlopok festése és tartományok elnevezése.
    
    Args:
    - file_path (str): Az Excel fájl elérési útja.
    - range_names (dict): A tartományok nevei az egyes munkalapokhoz.
    """
    wb = None
    try:
        wb = load_workbook(file_path)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            adjust_column_width(sheet)
            add_filter_and_paint(sheet)
            if sheet_name in range_names:
                name_ranges(sheet, range_names[sheet_name])
        wb.save(file_path)
    except Exception as e:
        logger.error(f"Hiba történt a munkafüzet feldolgozása közben: {e}")
    finally:
        if wb:
            wb.close()

def remove_illegal_characters(df):
    """
    Az illegális karakterek eltávolítása az összes sztring bejegyzésből a DataFrame-ben.
    
    Args:
    - df (pd.DataFrame): A tisztítandó DataFrame.
    
    Returns:
    - pd.DataFrame: A tisztított DataFrame.
    """
    # Illegális karakterek eltávolítása (ASCII tartomány: 0-31 és 127)
    illegal_characters = re.compile(r'[\x00-\x1F\x7F]')
    return df.applymap(lambda x: illegal_characters.sub('', x) if isinstance(x, str) else x)

def write_dataframes_to_excel(file_path, output_path, dataframes, sheet_names, start_cells):
    """
    Több DataFrame írása egy meglévő Excel fájlba, meghatározott celláktól kezdve.

    Args:
    - file_path (str): A meglévő Excel fájl elérési útja.
    - dataframes (list of pd.DataFrame): Az írandó DataFrame-ek listája.
    - sheet_names (list of str): A munkalapok neveinek listája, ahová a DataFrame-eket írjuk.
    - start_cells (list of str): A kezdő cellák listája (pl. 'I2') minden DataFrame-hez.
    """
    app = xw.App(visible=False)
    wb = app.books.open(file_path)
    
    try:
        for df, sheet_name, start_cell in zip(dataframes, sheet_names, start_cells):
            # Biztosítjuk, hogy a munkalap létezik
            if sheet_name not in [sheet.name for sheet in wb.sheets]:
                wb.sheets.add(sheet_name)
            
            # A munkalap lekérése
            sheet = wb.sheets[sheet_name]
            
            # A DataFrame írása a megadott helyre
            sheet.range(start_cell).options(index=False, header=False).value = df.values
        
        # A munkafüzet mentése
        wb.save(output_path)
    except Exception as e:
        print(f"Hiba történt a DataFrame-ek Excel fájlba írása közben: {e}")
    finally:
        wb.close()
        app.quit()